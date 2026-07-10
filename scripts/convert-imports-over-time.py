import json
import openpyxl

wb = openpyxl.load_workbook('data/CommodityImports.xlsx', data_only=True)
ws = wb.active

headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
year_cols = [(i, h) for i, h in enumerate(headers) if isinstance(h, int)]

out = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    iso3 = str(row[1]).strip() if row[1] else None
    if not iso3:
        continue
    series = []
    for col_idx, year in year_cols:
        val = row[col_idx]
        if val is not None and isinstance(val, (int, float)):
            series.append({'year': year, 'val': round(float(val) / 1000, 2)})
    if series:
        out[iso3] = series

out_path = 'public/assets/data/cdde_imports_over_time.json'
with open(out_path, 'w') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)
print(f'Written {len(out)} entries → {out_path}')
for iso3 in list(out.keys())[:3]:
    print(f'  {iso3}: {out[iso3][:3]}...')
