"""
Convert data/Treemap_CommodityExports.xlsx → public/assets/data/cdde_exports_by_region.csv

Excel structure:
  Col A: Region
  Col B: Sub region
  Col C: Commodity exports value (millions USD, 2022-2024 avg)
  Col D: Country name
  Col E: ISO code — integer (4-5 digits) for region/subregion aggregates,
                    3-char zero-padded string (e.g. '004') for individual countries

Only country rows (str ISO) are written to the CSV.
"""

import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess
    print("openpyxl not found — installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl",
                           "--break-system-packages", "-q"])
    import openpyxl

# Long UN country name fragments to strip for display
_STRIP_SUFFIXES = re.compile(
    r'\s*\((the|Plurinational State of|Bolivarian Republic of'
    r'|Islamic Republic of|Federated States of'
    r'|Kingdom of the|including Liechtenstein)\)\s*$',
    re.IGNORECASE,
)

_REPLACEMENTS = {
    "United Kingdom of Great Britain and Northern Ireland": "United Kingdom",
    "Democratic People's Republic of Korea": "DPR Korea",
    "Lao People's Democratic Republic": "Lao PDR",
    "Democratic Republic of the Congo": "DR Congo",
    "United Republic of Tanzania": "Tanzania",
    "State of Palestine": "Palestine",
    "Republic of Korea": "Republic of Korea",
    "Republic of Moldova": "Moldova",
    "Syrian Arab Republic": "Syria",
    "Cabo Verde": "Cape Verde",
    "Holy see": "Holy See",
}


def clean_name(name: str) -> str:
    name = name.strip()
    name = _STRIP_SUFFIXES.sub("", name)
    # Normalize Unicode apostrophes/quotes so dict lookups work regardless of source encoding
    name = name.replace("’", "'").replace("‘", "'")
    return _REPLACEMENTS.get(name, name)


def convert(xlsx_path: Path, csv_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows_written = 0
    totals_written = 0
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    # Collect subregion totals first (integer ISO, single subregion name without ';')
    subregion_totals = {}
    for row in ws.iter_rows(values_only=True):
        region, subregion, value, country, iso = row[:5]
        if not isinstance(iso, int):
            continue
        if region is None or subregion is None or value is None:
            continue
        if not isinstance(subregion, str) or ';' in subregion:
            continue
        key = (clean_name(str(region)), clean_name(str(subregion)))
        subregion_totals[key] = round(float(value))

    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["region", "subregion", "country", "value"])

        # Write subregion total rows first (empty country = total marker)
        for (region, subregion), total in subregion_totals.items():
            writer.writerow([region, subregion, "", total])
            totals_written += 1

        for row in ws.iter_rows(values_only=True):
            region, subregion, value, country, iso = row[:5]

            # Skip header, empty rows, and aggregate rows.
            # Country rows have a 3-char zero-padded string ISO code (e.g. '004');
            # aggregate rows have an integer ISO code (e.g. 5110).
            if not isinstance(iso, str):
                continue
            if region is None or country is None or value is None:
                continue
            if not isinstance(value, (int, float)):
                continue

            writer.writerow([
                clean_name(str(region)),
                clean_name(str(subregion)),
                clean_name(str(country)),
                round(float(value)),
            ])
            rows_written += 1

    print(f"Written {totals_written} subregion totals + {rows_written} country rows → {csv_path}")


if __name__ == "__main__":
    root = Path(__file__).parent.parent
    convert(
        root / "data" / "Treemap_CommodityExports.xlsx",
        root / "public" / "assets" / "data" / "cdde_exports_by_region.csv",
    )
