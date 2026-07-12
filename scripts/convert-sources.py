"""Convert data/Sources.xlsx into three JSON data files."""
import json
import os
from datetime import datetime
import openpyxl

XLSX = os.path.join(os.path.dirname(__file__), '..', 'data', 'Sources.xlsx')
TARGETS = [
    os.path.join(os.path.dirname(__file__), '..', 'public', 'assets', 'data'),
    os.path.join(os.path.dirname(__file__), '..', 'dist', 'assets', 'data'),
]

wb = openpyxl.load_workbook(XLSX)

# ── Sheet 1: Sources ────────────────────────────────────────────────────────

ws = wb['New Sources ']
rows = [tuple(c for c in row) for row in ws.iter_rows(values_only=True)]

sources = []
i = 0
while i < len(rows):
    a, b = rows[i][0], rows[i][1] if len(rows[i]) > 1 else None
    # Title row: non-empty A, B is None or empty
    if a and not b:
        title = str(a).strip()
        entry = {'title': title, 'unit': None, 'definition': None, 'comments': None, 'source': None, 'link': None, 'date': None}
        i += 1
        # Consume next rows until blank or next title
        while i < len(rows):
            ra, rb = rows[i][0], rows[i][1] if len(rows[i]) > 1 else None
            if ra is None and rb is None:
                break
            if ra and not rb and str(ra).strip() not in ('Parameter', 'Details'):
                break  # Next title block
            key = str(ra).strip().lower().rstrip() if ra else ''
            val = str(rb).strip() if rb else ''
            if key == 'unit':
                entry['unit'] = val
            elif key == 'definition':
                entry['definition'] = val
            elif key in ('comments', 'comments '):
                entry['comments'] = val
            elif key == 'source':
                entry['source'] = val
            elif key == 'link to the source':
                entry['link'] = val
            elif key == 'date extracted':
                if isinstance(rb, datetime):
                    entry['date'] = rb.strftime('%Y-%m-%d')
                else:
                    entry['date'] = str(rb).strip()
            i += 1
        # Fix Excel typo: second "Net Food Imports" block is actually Net Energy Imports
        if entry['title'] == 'Net Food Imports, 2012-2014 and 2022-2024' and entry.get('definition', '') and 'Energy' in entry['definition']:
            entry['title'] = 'Net Energy Imports, 2012-2014 and 2022-2024'
        # Clean up trailing whitespace in title
        entry['title'] = entry['title'].strip()
        sources.append(entry)
    else:
        i += 1

# ── Sheet 2: Country groupings ──────────────────────────────────────────────

ws2 = wb['Country groupings ']
rows2 = [tuple(c for c in row) for row in ws2.iter_rows(values_only=True)]

geographic = []
special = []

# Rows 2–20 (0-indexed 1–19): Region | Subregion | Countries
for row in rows2[1:20]:
    region = str(row[0]).strip() if row[0] else None
    subregion = str(row[1]).strip() if len(row) > 1 and row[1] else None
    countries_raw = str(row[2]).strip() if len(row) > 2 and row[2] else None
    if region and subregion and countries_raw:
        countries = [c.strip() for c in countries_raw.split(';') if c.strip()]
        geographic.append({'region': region, 'subregion': subregion, 'countries': sorted(countries)})

# Rows 23–28 (0-indexed 22–27): Label | Countries (semicolons in col B or col A contains both)
special_label_map = {
    'SIDS': 'Small Island Developing States (SIDS)',
    'LLDC': 'Land-Locked Developing Countries (LLDC)',
    'LDC': 'Least Developed Countries (LDC)',
    'Other developing countries ': 'Other developing countries',
    'Developed countries': 'Developed countries',
    'EU27': 'European Union (EU27)',
}
for row in rows2[22:28]:
    label_raw = str(row[0]).strip() if row[0] else None
    countries_raw = str(row[1]).strip() if len(row) > 1 and row[1] else None
    if label_raw and countries_raw:
        label = special_label_map.get(label_raw, label_raw.strip())
        sep = '; ' if '; ' in countries_raw else (', ' if ', ' in countries_raw else ';')
        countries = [c.strip() for c in countries_raw.split(sep) if c.strip()]
        special.append({'label': label, 'countries': sorted(countries)})

groupings = {'geographic': geographic, 'special': special}

# ── Sheet 3: Classification SITC 3 ─────────────────────────────────────────

ws3 = wb['Classification SITC 3']
rows3 = [row[0] for row in ws3.iter_rows(values_only=True)]

# Group by blank lines into sections. Skip title rows 1-2.
sections = []
current_heading = None
current_items = []

def parse_sitc_row(text):
    """Return (code, label) from '[code] label' or (None, text) for headings."""
    t = text.strip()
    if t.startswith('[') and ']' in t:
        bracket_end = t.index(']')
        code = t[1:bracket_end].strip()
        label = t[bracket_end + 1:].strip()
        return code, label
    return None, t

for raw in rows3[2:]:  # Skip rows 1-2 (title lines)
    if raw is None:
        # Blank line — flush current section
        if current_heading or current_items:
            sections.append({'heading': current_heading or '', 'items': current_items})
            current_heading = None
            current_items = []
        continue

    text = str(raw).strip()
    if not text:
        continue

    code, label = parse_sitc_row(text)
    if code is None:
        # No brackets — treat as a section heading or descriptive line
        if not current_items:
            current_heading = label
        else:
            # Descriptive text within a section (e.g. "All other products...")
            sections.append({'heading': current_heading or '', 'items': current_items})
            current_heading = label
            current_items = []
    else:
        # It's a classification item — the first item in a blank-separated group
        # may itself be the section heading (aggregate) if no prior heading set
        if current_heading is None and not current_items:
            current_heading = label  # Use the aggregate label as section heading
        current_items.append({'code': code, 'label': label})

# Flush last section
if current_heading or current_items:
    sections.append({'heading': current_heading or '', 'items': current_items})

# Filter out empty/note-only sections
sections = [s for s in sections if s['items'] and s['heading'] and not s['heading'].startswith('Note:')]
sitc3 = {'sections': sections}

# ── Write JSON ──────────────────────────────────────────────────────────────

def write_json(data, filename):
    for target in TARGETS:
        os.makedirs(target, exist_ok=True)
        path = os.path.join(target, filename)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f'Wrote {path}')

write_json(sources, 'cdde_sources.json')
write_json(groupings, 'cdde_commodity_groups.json')
write_json(sitc3, 'cdde_sitc3.json')

print(f'\nDone. {len(sources)} source entries, {len(geographic)} geographic groups, {len(special)} special categories, {len(sections)} SITC sections.')
