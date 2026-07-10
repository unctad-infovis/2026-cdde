import json
import openpyxl

def load_single(fname, key, col=4):
    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb.active
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        iso3 = str(row[1]).strip() if row[1] else None
        val = row[col]
        if iso3 and val is not None and isinstance(val, (int, float)):
            out[iso3] = val
    return out

emp = load_single('data/Employment.xlsx', 'emp')
emp_f = load_single('data/EmploymentFemane.xlsx', 'emp_female')
pop = load_single('data/Population.xlsx', 'population')

wb_hdi = openpyxl.load_workbook('data/HumanDevelopmentIndex.xlsx', data_only=True)
ws_hdi = wb_hdi.active
hdi_data = {}
for row in ws_hdi.iter_rows(min_row=2, values_only=True):
    iso3 = str(row[1]).strip() if row[1] else None
    if not iso3:
        continue
    rank, value, category = row[4], row[5], row[6]
    hdi_data[iso3] = {
        'hdi_rank': int(rank) if rank is not None and isinstance(rank, (int, float)) else None,
        'hdi_value': round(float(value), 3) if value is not None and isinstance(value, (int, float)) else None,
        'hdi_category': str(category).strip() if category else None,
    }

all_iso3 = set(emp) | set(emp_f) | set(pop) | set(hdi_data)
out = {}
for iso3 in sorted(all_iso3):
    entry = {}
    if iso3 in emp:
        entry['emp'] = round(float(emp[iso3]), 1)
    if iso3 in emp_f:
        entry['emp_female'] = round(float(emp_f[iso3]), 1)
    if iso3 in pop:
        entry['population'] = round(float(pop[iso3]))
    if iso3 in hdi_data:
        entry.update(hdi_data[iso3])
    out[iso3] = entry

out_path = 'public/assets/data/cdde_social_context.json'
with open(out_path, 'w') as f:
    json.dump(out, f, indent=2, ensure_ascii=False)
print(f'Written {len(out)} entries → {out_path}')
for iso3 in list(out.keys())[:3]:
    print(f'  {iso3}: {out[iso3]}')
